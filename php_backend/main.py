import tkinter as tk
from tkinter import ttk, messagebox
import mysql.connector
import cv2
import os
import numpy as np
from datetime import datetime, time
import time as time_module
import pyttsx3
import threading
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
import bcrypt

engine = pyttsx3.init()

# ---------- Path Configuration ----------
# Get the directory where this script is located
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
IMAGES_DIR = os.path.join(DATA_DIR, 'images')
TRAINING_DIR = os.path.join(DATA_DIR, 'training')
REPORTS_DIR = os.path.join(DATA_DIR, 'attendance_reports')

# Create directories if they don't exist
os.makedirs(IMAGES_DIR, exist_ok=True)
os.makedirs(TRAINING_DIR, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)

# ---------- Database Configuration ----------
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': 'root',
    'database': 'organization'
}

class AttendanceSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Employee Attendance System")
        self.recognizer = cv2.face.LBPHFaceRecognizer_create()
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        self.current_org_id = None
        self.org_name = ""
        self.org_email = ""
        self.last_voice_time = 0
        self.last_employee_announced = None
        self.training_in_progress = False
        self.show_login_screen()

    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def show_login_screen(self):
        self.clear_window()
        tk.Label(self.root, text="Organization Email:").pack(pady=5)
        self.email_entry = tk.Entry(self.root, width=30)
        self.email_entry.pack(pady=5)

        tk.Label(self.root, text="Password:").pack(pady=5)
        self.password_entry = tk.Entry(self.root, width=30, show="*")
        self.password_entry.pack(pady=5)

        tk.Button(self.root, text="Login", command=self.validate_login).pack(pady=20)

    def validate_login(self):
        email = self.email_entry.get()
        password = self.password_entry.get()

        try:
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT org_id, org_name, org_email, org_password FROM organization_register WHERE org_email=%s",
                (email,)
            )
            result = cursor.fetchone()

            if result:
                org_id, org_name, org_email, hashed_password = result
                if isinstance(hashed_password, str):
                    hashed_password = hashed_password.encode('utf-8')

                if bcrypt.checkpw(password.encode('utf-8'), hashed_password):
                    self.current_org_id = org_id
                    self.org_name = org_name
                    self.org_email = org_email
                    self.show_main_menu()
                    engine.say(f"Welcome to {self.org_name}")
                    engine.runAndWait()
                    self.init_attendance_excel()
                else:
                    messagebox.showerror("Error", "Invalid credentials")
            else:
                messagebox.showerror("Error", "Invalid credentials")

        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()

    def init_attendance_excel(self):
        current_date = datetime.now()
        year = current_date.year
        month = current_date.month
        month_name = calendar.month_name[month]

        filename = os.path.join(REPORTS_DIR, f"{self.org_name}_{month_name}_{year}_attendance.xlsx")

        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.title = f"{month_name} {year}"

            ws.merge_cells('A1:F1')
            ws['A1'] = f"Organization: {self.org_name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:F2')
            ws['A2'] = f"Month: {month_name}, Year: {year}"
            ws['A2'].font = Font(bold=True)
            ws['A2'].alignment = Alignment(horizontal='center')

            headers = ["Serial No", "Employee ID", "Name", "Status", "Timestamp", "Remarks"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header)
                ws.cell(row=3, column=col).font = Font(bold=True)

            days_in_month = calendar.monthrange(year, month)[1]
            for day in range(1, days_in_month + 1):
                col = 6 + day
                ws.cell(row=3, column=col, value=day)
                ws.cell(row=3, column=col).font = Font(bold=True)

            wb.save(filename)

            register_file = os.path.join(REPORTS_DIR, f"{self.org_name}_employee_register.xlsx")
            if not os.path.exists(register_file):
                wb_reg = Workbook()
                ws_reg = wb_reg.active
                ws_reg.title = "Employee Register"

                ws_reg.merge_cells('A1:E1')
                ws_reg['A1'] = f"Organization: {self.org_name}"
                ws_reg['A1'].font = Font(bold=True, size=14)
                ws_reg['A1'].alignment = Alignment(horizontal='center')

                reg_headers = ["Serial No", "Employee ID", "Name", "Join Date", "Category"]
                for col, header in enumerate(reg_headers, 1):
                    ws_reg.cell(row=2, column=col, value=header)
                    ws_reg.cell(row=2, column=col).font = Font(bold=True)

                wb_reg.save(register_file)

    def show_main_menu(self):
        self.clear_window()
        tk.Button(self.root, text="Add New Employee",
                  command=self.show_add_employee, height=2, width=20).pack(pady=10)
        tk.Button(self.root, text="Take Attendance",
                  command=self.take_attendance, height=2, width=20).pack(pady=10)
        tk.Button(self.root, text="Show Attendance",
                  command=self.show_attendance, height=2, width=20).pack(pady=10)
        tk.Button(self.root, text="Generate Report",
                  command=self.generate_report, height=2, width=20).pack(pady=10)
        tk.Button(self.root, text="Generate Wage Muster",
                  command=self.generate_wage_muster, height=2, width=20).pack(pady=10)
        tk.Button(self.root, text="Logout",
                  command=self.show_login_screen, height=2, width=20).pack(pady=10)

    def show_add_employee(self):
        self.clear_window()
        tk.Label(self.root, text="Employee ID:").pack(pady=5)
        self.emp_id_entry = tk.Entry(self.root)
        self.emp_id_entry.pack(pady=5)

        tk.Label(self.root, text="Employee Name:").pack(pady=5)
        self.emp_name_entry = tk.Entry(self.root)
        self.emp_name_entry.pack(pady=5)

        tk.Button(self.root, text="Verify", command=self.verify_employee).pack(pady=10)
        self.verify_status = tk.Label(self.root, text="")
        self.verify_status.pack(pady=5)

        self.take_img_btn = tk.Button(self.root, text="Take Image",
                                       command=self.capture_images, state=tk.DISABLED)
        self.take_img_btn.pack(pady=5)

        self.train_img_btn = tk.Button(self.root, text="Train Image",
                                        command=self.start_training, state=tk.DISABLED)
        self.train_img_btn.pack(pady=5)

        tk.Button(self.root, text="Back", command=self.show_main_menu).pack(pady=10)

    def verify_employee(self):
        emp_id = self.emp_id_entry.get()
        emp_name = self.emp_name_entry.get()

        if not emp_id or not emp_name:
            self.verify_status.config(text="Please enter both ID and Name", fg="red")
            engine.say("Please enter both employee ID and name")
            engine.runAndWait()
            return

        try:
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()
            cursor.execute("SELECT emp_id, empname, category FROM employees WHERE emp_id = %s AND empname = %s AND org_id = %s",
                           (emp_id, emp_name, self.current_org_id))
            result = cursor.fetchone()

            if result:
                self.verify_status.config(text="Verified", fg="green")
                self.take_img_btn.config(state=tk.NORMAL)
                self.emp_id = emp_id
                self.emp_name = emp_name
                self.emp_category = result[2]
                engine.say(f"Employee {emp_name} verified. You can now capture images.")
                engine.runAndWait()

                register_file = os.path.join(REPORTS_DIR, f"{self.org_name}_employee_register.xlsx")
                if os.path.exists(register_file):
                    wb = load_workbook(register_file)
                    ws = wb.active

                    for row in range(3, ws.max_row + 1):
                        if ws.cell(row=row, column=2).value == int(emp_id):
                            self.verify_status.config(text="Employee already registered", fg="red")
                            engine.say("Employee already registered")
                            engine.runAndWait()
                            self.take_img_btn.config(state=tk.DISABLED)
                            wb.close()
                            return
                    wb.close()

                # Create employee image folder inside data/images
                emp_img_dir = os.path.join(IMAGES_DIR, str(emp_id))
                os.makedirs(emp_img_dir, exist_ok=True)
            else:
                self.verify_status.config(text="Not registered in this organization", fg="red")
                engine.say("Employee not registered in this organization")
                engine.runAndWait()
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()

    def capture_images(self):
        engine.say("Starting image capture. Please position your face in the center of the frame.")
        engine.runAndWait()

        cam = cv2.VideoCapture(0)
        if not cam.isOpened():
            engine.say("Camera not detected. Please check your camera connection.")
            engine.runAndWait()
            return

        count = 0
        emp_img_dir = os.path.join(IMAGES_DIR, str(self.emp_id))

        while True:
            ret, img = cam.read()
            if not ret:
                engine.say("Camera error. Please check your camera connection.")
                engine.runAndWait()
                break

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                cv2.rectangle(img, (x, y), (x+w, y+h), (255, 0, 0), 2)
                count += 1
                cv2.imwrite(os.path.join(emp_img_dir, f"image_{count}.jpg"), gray[y:y+h, x:x+w])

            cv2.imshow(f"Capture Images for {self.emp_name} - Press SPACE when ready, ESC to exit", img)

            k = cv2.waitKey(100) & 0xff
            if k == 27:
                break
            elif k == 32 or count >= 30:
                break

        cam.release()
        cv2.destroyAllWindows()

        if count > 0:
            self.train_img_btn.config(state=tk.NORMAL)
            engine.say(f"{count} images captured successfully for {self.emp_name}. You can now train the model.")
            engine.runAndWait()
        else:
            engine.say("No images captured. Please try again.")
            engine.runAndWait()

    def start_training(self):
        if self.training_in_progress:
            return

        self.training_in_progress = True
        self.train_img_btn.config(state=tk.DISABLED)

        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Training Progress")
        self.progress_window.geometry("400x150")
        self.progress_window.resizable(False, False)

        tk.Label(self.progress_window, text="Training in progress...", font=("Arial", 14)).pack(pady=10)

        self.progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(self.progress_window, variable=self.progress_var, maximum=100, length=300)
        progress_bar.pack(pady=10)

        self.percentage_label = tk.Label(self.progress_window, text="0%", font=("Arial", 12))
        self.percentage_label.pack()

        self.status_label = tk.Label(self.progress_window, text="Initializing...", font=("Arial", 10))
        self.status_label.pack(pady=5)

        threading.Thread(target=self.train_images, daemon=True).start()

    def update_progress(self, value, message):
        self.progress_var.set(value)
        self.percentage_label.config(text=f"{int(value)}%")
        self.status_label.config(text=message)
        self.progress_window.update_idletasks()

    def train_images(self):
        try:
            faces = []
            ids = []
            image_paths = []

            for emp_id in os.listdir(IMAGES_DIR):
                if not emp_id.isdigit():
                    continue
                emp_folder = os.path.join(IMAGES_DIR, emp_id)
                if not os.path.isdir(emp_folder):
                    continue
                for image_path in os.listdir(emp_folder):
                    full_path = os.path.join(emp_folder, image_path)
                    image_paths.append((emp_id, full_path))

            total_images = len(image_paths)
            if total_images == 0:
                self.update_progress(100, "No images found for training")
                engine.say("No images found for training")
                engine.runAndWait()
                time_module.sleep(2)
                self.progress_window.destroy()
                self.training_in_progress = False
                return

            processed = 0
            for emp_id, img_path in image_paths:
                img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
                if img is not None:
                    faces.append(img)
                    ids.append(int(emp_id))

                processed += 1
                progress = (processed / total_images) * 100
                self.update_progress(progress, f"Processing image {processed}/{total_images}")
                time_module.sleep(0.05)

            self.update_progress(95, "Training model...")
            if faces:
                self.recognizer.train(faces, np.array(ids))
                training_file = os.path.join(TRAINING_DIR, 'training_data.yml')
                self.recognizer.save(training_file)
                self.update_progress(100, "Training complete!")

                register_file = os.path.join(REPORTS_DIR, f"{self.org_name}_employee_register.xlsx")
                wb = load_workbook(register_file)
                ws = wb.active

                next_row = ws.max_row + 1
                ws.cell(row=next_row, column=1, value=next_row-2)  # Serial No
                ws.cell(row=next_row, column=2, value=int(self.emp_id))  # Employee ID
                ws.cell(row=next_row, column=3, value=self.emp_name)  # Name
                ws.cell(row=next_row, column=4, value=datetime.now().strftime("%Y-%m-%d"))  # Join Date
                ws.cell(row=next_row, column=5, value=self.emp_category)  # Category

                wb.save(register_file)
                wb.close()

                engine.say("Training completed successfully")
                engine.runAndWait()
                time_module.sleep(1)
                messagebox.showinfo("Success", "Employee registered and trained successfully")
            else:
                self.update_progress(100, "Training failed: no valid images")
                engine.say("Training failed. No valid images found")
                engine.runAndWait()
                time_module.sleep(1)
                messagebox.showwarning("Training Error", "No valid images found for training")
        except Exception as e:
            self.update_progress(100, f"Error: {str(e)}")
            engine.say(f"Training error: {str(e)}")
            engine.runAndWait()
            time_module.sleep(1)
            messagebox.showerror("Training Error", str(e))
        finally:
            time_module.sleep(1)
            self.progress_window.destroy()
            self.training_in_progress = False

    def generate_wage_muster(self):
        try:
            import webbrowser
            import subprocess
            import sys
            import socket

            wage_muster_dir = os.path.join(BASE_DIR, 'wage_muster_generator')

            if not os.path.exists(wage_muster_dir):
                messagebox.showerror(
                    "Error",
                    "Wage Muster Generator folder not found.\n\n"
                    "Please create 'wage_muster_generator' folder with all required files."
                )
                engine.say("Wage muster generator folder not found")
                engine.runAndWait()
                return

            app_py_path = os.path.join(wage_muster_dir, 'app.py')
            if not os.path.exists(app_py_path):
                messagebox.showerror(
                    "Error",
                    "Wage Muster Generator app.py not found.\n\n"
                    "Please place app.py in wage_muster_generator folder."
                )
                engine.say("Wage muster generator application not found")
                engine.runAndWait()
                return

            run_script_path = os.path.join(wage_muster_dir, 'run_flask.py')

            if not os.path.exists(run_script_path):
                run_script_content = '''"""
Simple script to run the Flask wage muster generator
"""
import os
import sys

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app

if __name__ == '__main__':
    # Clean up old files before starting
    app.cleanup_old_files()
    print("Starting Wage Muster Generator on http://127.0.0.1:5000")
    print("Press Ctrl+C to stop the server")
    app.run(debug=False, host='127.0.0.1', port=5000, use_reloader=False)
'''
                with open(run_script_path, 'w') as f:
                    f.write(run_script_content)

            def is_port_in_use(port):
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    return s.connect_ex(('localhost', port)) == 0

            if not is_port_in_use(5000):
                print("Starting Flask server for Wage Muster Generator...")

                if sys.platform == 'win32':
                    startupinfo = subprocess.STARTUPINFO()
                    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                    startupinfo.wShowWindow = subprocess.SW_HIDE

                    flask_process = subprocess.Popen(
                        [sys.executable, run_script_path],
                        cwd=wage_muster_dir,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        startupinfo=startupinfo,
                        creationflags=subprocess.CREATE_NO_WINDOW
                    )
                else:
                    flask_process = subprocess.Popen(
                        [sys.executable, run_script_path],
                        cwd=wage_muster_dir,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        start_new_session=True
                    )

                self.flask_process = flask_process
                time_module.sleep(3)

                if is_port_in_use(5000):
                    messagebox.showinfo(
                        "Wage Muster Generator",
                        "Wage Muster Generator opened in your browser.\n\n"
                        "Note: The Flask server will run in background.\n"
                        "Close this window when done."
                    )
                else:
                    try:
                        flask_process.terminate()
                    except:
                        pass

                    messagebox.showerror(
                        "Error",
                        "Could not start Wage Muster Generator.\n"
                        "Port 5000 might be in use or there's an error in the Flask app."
                    )
                    engine.say("Failed to start wage muster generator")
                    engine.runAndWait()
                    return

            url = "http://127.0.0.1:5000"
            webbrowser.open(url)

            engine.say("Opening wage muster generator in your browser")
            engine.runAndWait()

        except Exception as e:
            error_msg = f"Error opening wage muster generator: {str(e)}"
            messagebox.showerror("Error", error_msg)
            engine.say("Error opening wage muster generator")
            engine.runAndWait()

    def take_attendance(self):
        training_file = os.path.join(TRAINING_DIR, 'training_data.yml')
        if not os.path.exists(training_file):
            messagebox.showwarning("Error", "Train images first!")
            engine.say("Please train images first before taking attendance")
            engine.runAndWait()
            return

        try:
            self.recognizer.read(training_file)
        except:
            messagebox.showwarning("Error", "Training data corrupted. Please retrain images.")
            engine.say("Training data corrupted. Please retrain images.")
            engine.runAndWait()
            return

        engine.say("Starting attendance system. Please face the camera.")
        engine.runAndWait()

        cam = cv2.VideoCapture(0)
        if not cam.isOpened():
            engine.say("Camera not detected. Please check your camera connection.")
            engine.runAndWait()
            return

        font = cv2.FONT_HERSHEY_SIMPLEX

        register_file = os.path.join(REPORTS_DIR, f"{self.org_name}_employee_register.xlsx")
        if not os.path.exists(register_file):
            messagebox.showwarning(
                "No Employees",
                "No employees registered yet. Please add employees first."
            )
            engine.say("No employees found. Please add employees first.")
            engine.runAndWait()
            return

        wb_reg = load_workbook(register_file)
        ws_reg = wb_reg.active
        registered_employees = {}

        for row in range(3, ws_reg.max_row + 1):
            emp_id = ws_reg.cell(row=row, column=2).value
            emp_name = ws_reg.cell(row=row, column=3).value
            if emp_id and emp_name:
                registered_employees[str(emp_id)] = emp_name

        wb_reg.close()

        if not registered_employees:
            messagebox.showwarning("No Employees", "Employee register is empty. Please add employees first.")
            engine.say("No employees in register. Please add employees first.")
            engine.runAndWait()
            return

        current_date = datetime.now()
        year = current_date.year
        month = current_date.month
        month_name = calendar.month_name[month]
        day = current_date.day

        attendance_file = os.path.join(REPORTS_DIR, f"{self.org_name}_{month_name}_{year}_attendance.xlsx")
        if not os.path.exists(attendance_file):
            self.init_attendance_excel()

        wb_att = load_workbook(attendance_file)
        ws_att = wb_att.active

        marked_today = set()
        for row in range(4, ws_att.max_row + 1):
            if ws_att.cell(row=row, column=5).value and ws_att.cell(row=row, column=5).value.date() == current_date.date():
                marked_today.add(str(ws_att.cell(row=row, column=2).value))

        self.last_voice_time = time_module.time()
        self.last_employee_announced = None

        while True:
            ret, img = cam.read()
            if not ret:
                continue

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray, 1.3, 5)
            for (x, y, w, h) in faces:
                cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)
                id, confidence = self.recognizer.predict(gray[y:y+h, x:x+w])

                if confidence < 50:
                    emp_id = str(id)
                    if emp_id in registered_employees:
                        name = registered_employees[emp_id]

                        cv2.putText(img, f"ID: {emp_id}", (x, y-40), font, 0.8, (0, 255, 0), 2)
                        cv2.putText(img, f"Name: {name}", (x, y-10), font, 0.8, (0, 255, 0), 2)

                        if emp_id in marked_today:
                            cv2.putText(img, "Already Marked", (x, y+h+20), font, 0.8, (0, 0, 255), 2)

                            current_time = time_module.time()
                            if (current_time - self.last_voice_time > 3 or
                                self.last_employee_announced != emp_id):
                                engine.say(f"{name}, you are already marked for today")
                                engine.runAndWait()
                                self.last_voice_time = current_time
                                self.last_employee_announced = emp_id
                        else:
                            now = datetime.now()
                            current_time_val = now.time()

                            if current_time_val < time(8, 0, 0):
                                status = "P"
                                db_status = "present"
                                remarks = "Present before 8:00 AM"
                            elif current_time_val < time(9, 0, 0):
                                status = "LM"
                                db_status = "late"
                                remarks = "Present between 8:00-9:00 AM"
                            elif current_time_val < time(12, 0, 0):
                                status = "A"
                                db_status = "absent"
                                remarks = "Present after 12:00 PM"
                            else:
                                status = "A"
                                db_status = "absent"
                                remarks = "Present after 12:00 PM"

                            # Mark attendance in Excel
                            next_row = ws_att.max_row + 1
                            ws_att.cell(row=next_row, column=1, value=next_row-3)
                            ws_att.cell(row=next_row, column=2, value=int(emp_id))
                            ws_att.cell(row=next_row, column=3, value=name)
                            ws_att.cell(row=next_row, column=4, value=status)
                            ws_att.cell(row=next_row, column=5, value=now)
                            ws_att.cell(row=next_row, column=6, value=remarks)

                            date_col = 6 + day
                            ws_att.cell(row=next_row, column=date_col, value=status)

                            # Insert into database
                            try:
                                conn = mysql.connector.connect(**db_config)
                                cursor = conn.cursor()
                                insert_query = """
                                    INSERT INTO attendance 
                                    (emp_id, org_id, date, status, time_in, created_at)
                                    VALUES (%s, %s, %s, %s, %s, NOW())
                                """
                                cursor.execute(insert_query,
                                              (int(emp_id), self.current_org_id, now.date(), db_status, now.time()))
                                conn.commit()
                                cursor.close()
                                conn.close()
                            except mysql.connector.IntegrityError as ie:
                                print(f"Integrity error (duplicate): {ie}")
                            except Exception as e:
                                print(f"Database insert error: {e}")
                                messagebox.showwarning(
                                    "Database Warning",
                                    f"Attendance marked in Excel but database insert failed: {e}"
                                )

                            marked_today.add(emp_id)

                            engine.say(f"Attendance marked for {name}. Status: {status}")
                            engine.runAndWait()
                            self.last_employee_announced = emp_id
                            self.last_voice_time = time_module.time()
                    else:
                        cv2.putText(img, "Employee Not Registered", (x, y+h+20), font, 0.8, (0, 0, 255), 2)

                        current_time = time_module.time()
                        if current_time - self.last_voice_time > 3:
                            engine.say("Employee not registered. Please contact administrator")
                            engine.runAndWait()
                            self.last_voice_time = current_time
                else:
                    cv2.putText(img, "Unknown", (x, y-10), font, 0.8, (0, 0, 255), 2)

                    current_time = time_module.time()
                    if current_time - self.last_voice_time > 3:
                        engine.say("Face not recognized. Please try again")
                        engine.runAndWait()
                        self.last_voice_time = current_time

            cv2.imshow("Take Attendance - Press ESC to exit", img)
            if cv2.waitKey(1) == 27:
                wb_att.save(attendance_file)
                wb_att.close()
                engine.say("Attendance session ended")
                engine.runAndWait()
                break

        cam.release()
        cv2.destroyAllWindows()

    def show_attendance(self):
        self.clear_window()
        engine.say("Showing attendance records")
        engine.runAndWait()

        try:
            current_date = datetime.now()
            year = current_date.year
            month = current_date.month
            month_name = calendar.month_name[month]

            attendance_file = os.path.join(REPORTS_DIR, f"{self.org_name}_{month_name}_{year}_attendance.xlsx")

            if os.path.exists(attendance_file):
                wb = load_workbook(attendance_file)
                ws = wb.active

                tree_frame = tk.Frame(self.root)
                tree_frame.pack(fill='both', expand=True, padx=10, pady=10)

                columns = []
                for col in range(1, ws.max_column + 1):
                    columns.append(ws.cell(row=3, column=col).value or f"Col{col}")

                tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
                scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
                tree.configure(yscrollcommand=scrollbar.set)
                scrollbar.pack(side="right", fill="y")

                for col, col_name in enumerate(columns, 1):
                    tree.heading(f'#{col}', text=col_name)
                    tree.column(f'#{col}', width=100, anchor='center')

                for row in range(4, ws.max_row + 1):
                    values = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        if isinstance(cell_value, datetime):
                            values.append(cell_value.strftime("%Y-%m-%d %H:%M"))
                        else:
                            values.append(cell_value)
                    tree.insert('', 'end', values=values)

                tree.pack(side='left', fill='both', expand=True)
                wb.close()
            else:
                tk.Label(self.root, text="No attendance records found", font=("Arial", 14)).pack(pady=20)
        except Exception as e:
            messagebox.showerror("Error", f"Couldn't load attendance: {str(e)}")

        tk.Button(self.root, text="Back", command=self.show_main_menu, width=15).pack(pady=10)

    def generate_report(self):
        current_date = datetime.now()
        year = current_date.year
        month = current_date.month
        month_name = calendar.month_name[month]
        day = current_date.day

        attendance_file = os.path.join(REPORTS_DIR, f"{self.org_name}_{month_name}_{year}_attendance.xlsx")
        register_file = os.path.join(REPORTS_DIR, f"{self.org_name}_employee_register.xlsx")

        if not os.path.exists(attendance_file) or not os.path.exists(register_file):
            messagebox.showwarning("Error", "Attendance or register file not found!")
            return

        wb_att = load_workbook(attendance_file)
        ws_att = wb_att.active
        wb_reg = load_workbook(register_file)
        ws_reg = wb_reg.active

        registered_employees = {}
        for row in range(3, ws_reg.max_row + 1):
            emp_id = ws_reg.cell(row=row, column=2).value
            emp_name = ws_reg.cell(row=row, column=3).value
            if emp_id and emp_name:
                registered_employees[str(emp_id)] = emp_name

        marked_today = set()
        for row in range(4, ws_att.max_row + 1):
            if ws_att.cell(row=row, column=5).value and ws_att.cell(row=row, column=5).value.date() == current_date.date():
                marked_today.add(str(ws_att.cell(row=row, column=2).value))

        for emp_id, emp_name in registered_employees.items():
            if emp_id not in marked_today:
                next_row = ws_att.max_row + 1
                ws_att.cell(row=next_row, column=1, value=next_row-3)
                ws_att.cell(row=next_row, column=2, value=int(emp_id))
                ws_att.cell(row=next_row, column=3, value=emp_name)
                ws_att.cell(row=next_row, column=4, value="A")
                ws_att.cell(row=next_row, column=5, value=current_date)
                ws_att.cell(row=next_row, column=6, value="Absent - Not marked")

                date_col = 6 + day
                ws_att.cell(row=next_row, column=date_col, value="A")

        wb_att.save(attendance_file)
        wb_att.close()
        wb_reg.close()

        messagebox.showinfo("Success", "Attendance report generated successfully!")
        engine.say("Attendance report generated successfully")
        engine.runAndWait()

    def cleanup(self):
        if hasattr(self, 'flask_process'):
            try:
                self.flask_process.terminate()
                time_module.sleep(1)
                if self.flask_process.poll() is None:
                    self.flask_process.kill()
            except:
                pass

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("600x500")
    root.title("Employee Attendance System")
    app = AttendanceSystem(root)

    def on_closing():
        app.cleanup()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()